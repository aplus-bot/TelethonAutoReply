"""
A combined Telethon bot that provides both an auto-responder menu and an advanced invoice scanning/reporting system.
"""

import logging
import asyncio
import re
import os
import io
from datetime import datetime, timedelta
from collections import defaultdict
from telethon import TelegramClient, events
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
from telethon.sessions import StringSession

# ====== LOGGING ======
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.WARNING
)

# ====== CONFIGURATION ======
def read_config(file_path):
    config = {}
    try:
        with open(file_path, "r") as f:
            for line in f:
                if "=" in line:
                    key, value = line.split("=", 1)
                    value = value.split("#")[0].strip().strip("'").strip('"')
                    config[key.strip()] = value
    except FileNotFoundError:
        print(f"⚠️រកមិនឃើញឯកសារ {file_path}")
    return config

# Load API Config
config = read_config("TelethonAPI.txt")
API_ID = int(os.environ.get("API_ID") or config.get("API_ID", 0))
API_HASH = os.environ.get("API_HASH") or config.get("API_HASH", "")
BOT_TOKEN = os.environ.get("BOT_TOKEN") or config.get("BOT_TOKEN", "")
SESSION_STRING = os.environ.get("SESSION_STRING", "")

# --- Auto Reply Config ---
SOURCE_CHANNEL = "@Aplus_V5"
MENU_MSG_ID = 3
CONTENT_MAP = {
    "1": [4],
    "2": [6, 7, 8, 9, 12],
    "3": [13, 14, 15, 16],
    "4": [17],
    "5": [18, 19, 28, 29, 30, 31],
    "6": [32, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 49],
    "7": [50]
}

# --- Invoice Bot Config ---
MAX_INVOICES = 100

# Initialize Client
if SESSION_STRING:
    client = TelegramClient(StringSession(SESSION_STRING), API_ID, API_HASH)
else:
    client = TelegramClient('combined_bot_session', API_ID, API_HASH)

# ====== PATTERNS ======
# Invoice Patterns
invoice_pattern = re.compile(
    r"(?:🧾|🧾\s*(?:វិក្កយបត្រ|Receipt|RECEIPT))\s*(\d+).*?"  # Group 1: Invoice No
    r"(?:💵\s*(?:សរុប|Total)\s*:\s*\$([\d,]+\.?\d*)\s*\|\s*R\.?\s*([\d,]+))|"  # Group 2,3: USD, Riel (Format 1)
    r"(?:💵\s*(?:សរុប|Total)\s*:\s*\$([\d,]+\.?\d*)\s*\|\s*R\.?\s*([\d,]+)).*?"  # Group 4,5: USD, Riel (Format 2)
    r"(?:📆|📅)\s*(\d{1,2}-[A-Za-z]{3}-\d{4})",  # Group 6: Date (DD-MMM-YYYY)
    re.DOTALL | re.IGNORECASE
)

alt_invoice_pattern = re.compile(
    r"(?:🧾|🧾\s*(?:វិក្កយបត្រ|Receipt|RECEIPT))\s*(\d+).*?"  # Group 1: Invoice No
    r"(?:📆|📅)\s*(\d{1,2}/?\d{0,2}/?\d{4}).*?"  # Group 2: Date (DD/MM/YYYY)
    r"(?:💵\s*(?:សរុប|Total)\s*:\s*\$([\d,]+\.?\d*)\s*\|\s*R\.?\s*([\d,]+))",  # Group 3,4: USD, Riel
    re.DOTALL | re.IGNORECASE
)

# Patterns for extra details (Student ID, Name, Phone)
student_id_pattern = re.compile(r"(?:✅|🔑)?\s*(?:កូដសិស្ស|Stu\.?ID|StudentCode)\s*:\s*(\S+)", re.IGNORECASE)
name_pattern = re.compile(r"(?:✅\s*(?:គោត្តនាម-នាម|Full Name|Name|ឈ្មោះ)\s*:\s*|👤\s*)(.+?)(?=\s*(?:✅|📞|Phone|ទូរស័ព្ទ|🎯|--|$)|\n)", re.IGNORECASE)
phone_pattern = re.compile(r"(?:✅|📞)\s*(?:ទូរស័ព្ទ|Phone)\s*:\s*([\d\s]+)", re.IGNORECASE)

# ====== HELPER FUNCTIONS ======
def parse_date(date_str):
    months = {
        'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
        'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
    }
    try:
        if '-' in date_str and any(month in date_str.lower() for month in months.keys()):
            day, month, year = date_str.split('-')
            return datetime(int(year), months[month.lower()[:3]], int(day)).date()
        elif '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 3:
                day, month, year = parts
                return datetime(int(year), int(month), int(day)).date()
    except:
        return None

async def get_invoices(chat, start_date, end_date=None):
    """Scans chat history for invoices within a given date range."""
    if end_date is None:
        end_date = start_date
    invoices = []
    seen_invoices = set()
    
    # 3. Smart Date Break: Stop scanning if messages are older than start_date - 10 days
    cutoff_date = start_date - timedelta(days=10)
    
    async for msg in client.iter_messages(chat, limit=None):
        if msg.date.date() < cutoff_date:
            break
            
        try:
            if msg.text:
                match = invoice_pattern.search(msg.text)
                if match:
                    invoice_no = match.group(1)
                    usd = match.group(2) or match.group(4)
                    riel = match.group(3) or match.group(5)
                    date_str = match.group(6) if match.group(6) else None
                    
                    if not date_str:
                        date_match = re.search(r'(?:📆|📅)\s*(\d{1,2}-[A-Za-z]{3}-\d{4})', msg.text)
                        if date_match:
                            date_str = date_match.group(1)
                else:
                    match = alt_invoice_pattern.search(msg.text)
                    if match:
                        invoice_no = match.group(1)
                        date_str = match.group(2)
                        usd = match.group(3)
                        riel = match.group(4)
                
                # Prevent Duplicates
                if match and invoice_no in seen_invoices:
                    continue
                
                # Extract extra details (ID, Name, Phone)
                s_id, s_name, s_phone = "", "", ""
                
                id_match = student_id_pattern.search(msg.text)
                if id_match: s_id = id_match.group(1).strip()
                
                name_match = name_pattern.search(msg.text)
                if name_match: s_name = name_match.group(1).strip()
                
                # Fallback: If ID found but Name not found via pattern, check the line below ID
                if s_id and not s_name and id_match:
                    # Get text after the ID
                    remaining_text = msg.text[id_match.end():]
                    # Split by newline and find the first non-empty line/segment
                    parts = [p.strip() for p in remaining_text.split('\n') if p.strip()]
                    if parts:
                        # Take the first part as name, provided it doesn't look like a phone number or total
                        candidate = parts[0]
                        if not any(k in candidate.lower() for k in ['phone', 'ទូរស័ព្ទ', 'total', 'សរុប']):
                            # Remove simple prefixes like "Name:" or ":" if present
                            s_name = re.sub(r'^(?:Name|ឈ្មោះ|:)\s*', '', candidate, flags=re.IGNORECASE).strip()
                
                phone_match = phone_pattern.search(msg.text)
                if phone_match: s_phone = phone_match.group(1).strip()
                
                if match and invoice_no and usd and riel:
                    if inv_date := parse_date(date_str) if date_str else msg.date.date():
                        if start_date <= inv_date <= end_date:
                            seen_invoices.add(invoice_no)
                            invoices.append({
                                'no': invoice_no.strip(),
                                'usd': float(usd.replace(",", "")),
                                'riel': int(riel.replace(",", "")),
                                'date': inv_date,
                                'msg_date': msg.date,
                                'student_id': s_id,
                                'student_name': s_name,
                                'phone': s_phone
                            })
        except Exception as e:
            logging.error(f"Error processing message: {e}")
    return invoices

async def send_report(event, invoices, period_name):
    """Generates and sends Excel and PDF reports."""
    if not invoices:
        await event.reply(f"📭 No invoices found for {period_name}")
        return
    
    invoices.sort(key=lambda x: int(x['no']))
    
    total_usd = sum(i['usd'] for i in invoices)
    total_riel = sum(i['riel'] for i in invoices)
    
    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Headers & Styling
    headers = ['Invoice No', 'Date', 'ID', 'Student Name', 'Phone', 'USD', 'Riel']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data Rows
    for r, inv in enumerate(invoices, 2):
        ws.cell(row=r, column=1, value=inv['no'].zfill(6))
        ws.cell(row=r, column=2, value=inv['date'])
        ws.cell(row=r, column=3, value=inv['student_id'])
        ws.cell(row=r, column=4, value=inv['student_name'])
        ws.cell(row=r, column=5, value=inv['phone'])
        
        # Currency Formatting
        usd_cell = ws.cell(row=r, column=6, value=inv['usd'])
        usd_cell.number_format = '"$"#,##0.00_-'
        
        riel_cell = ws.cell(row=r, column=7, value=inv['riel'])
        riel_cell.number_format = '#,##0 "៛"'

    # Total Row
    last_row = len(invoices) + 1
    total_row_idx = last_row + 1
    
    ws.cell(row=total_row_idx, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row_idx, column=6, value=total_usd).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=6).number_format = '"$"#,##0.00_-'
    ws.cell(row=total_row_idx, column=7, value=total_riel).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=7).number_format = '#,##0 "៛"'

    # Save to memory
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    safe_name = re.sub(r'[\\/*?:"<>|]', "", period_name).replace(" ", "_")
    file_bytes.name = f"{safe_name}.xlsx"

    # ====== CREATE PDF REPORT (Using FPDF2) ======
    pdf = FPDF()
    pdf.add_page()
    
    # 1. Setup Khmer Font
    font_path = 'KhmerOScontent.ttf' if os.path.exists('KhmerOScontent.ttf') else r'C:\Windows\Fonts\daunpenh.ttf'
    try:
        pdf.add_font("Khmer", fname=font_path)
        pdf.set_font("Khmer", size=10)
        # Enable Text Shaping (Crucial for Khmer Subscripts/Legs)
        pdf.set_text_shaping(True)
    except Exception as e:
        print(f"⚠️ Font Error: {e}")
        pdf.set_font("Helvetica", size=10)

    # 2. Title
    pdf.set_font(size=14)
    pdf.cell(0, 10, f"Report: {period_name}", align='C', new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # 3. Prepare Table Data
    pdf.set_font(size=9)
    table_data = [['No', 'Date', 'ID', 'Name', 'Phone', 'USD', 'Riel']]
    
    for inv in invoices:
        table_data.append([
            inv['no'], 
            inv['date'].strftime('%d-%b'), 
            inv['student_id'],
            inv['student_name'], 
            inv['phone'], 
            f"${inv['usd']:,.2f}", 
            f"{inv['riel']:,}"
        ])
    
    table_data.append(['TOTAL', '', '', '', '', f"${total_usd:,.2f}", f"{total_riel:,}"])

    # 4. Render Table
    with pdf.table() as table:
        for data_row in table_data:
            row = table.row()
            for datum in data_row:
                row.cell(datum)

    # 5. Output PDF
    pdf_buffer = io.BytesIO(pdf.output())
    pdf_buffer.seek(0)
    pdf_buffer.name = f"{safe_name}.pdf"
    
    summary = (
        f"📊 Report: {period_name}\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"💰 Total USD: ${total_usd:,.2f}\n"
        f"💰 Total Riel: {total_riel:,} ៛\n"
        f"🧾 Total Invoices: {len(invoices)}\n"
        f"📂 Reports attached (Excel & PDF)."
    )
    
    await event.client.send_file(event.chat_id, file=[pdf_buffer, file_bytes], caption=summary)

async def send_content_by_selection(event, selection):
    """Handles the logic for the auto-reply menu selections."""
    chat_id = event.chat_id
    message_ids = CONTENT_MAP.get(selection, [])
    
    if not message_ids:
        text = "👨‍💼 សូមទាក់ទង លោក ង៉ូវ សំណាង តាមរយៈតេឡេក្រាម។" if selection == "7" else "ℹ️ ព័ត៌មាននេះកំពុងរៀបចំ។"
        await client.send_message(chat_id, text)
        await client.send_message(chat_id, " វាយលេខ ០ (Back to Menu)")
        return

    status_msg = await client.send_message(chat_id, f"⏳ កំពុងផ្ញើជូនព័ត៌មានផ្នែកទី {selection}...")

    try:
        ids = [int(mid) for mid in message_ids]
        messages = await client.get_messages(SOURCE_CHANNEL, ids=ids)
        
        if not isinstance(messages, list): messages = [messages]
        messages = [m for m in messages if m is not None]

        for i, msg in enumerate(messages):
            await client.send_message(chat_id, msg)
            await asyncio.sleep(0.3)
        
        await client.send_message(chat_id, "💬 វាយលេខ ០ (Back to Menu)")
                
    except Exception as e:
        await client.send_message(chat_id, "❌ មានបញ្ហាក្នុងការទាញយកទិន្នន័យ។")
        logging.error(f"Error: {e}")
    finally:
        await status_msg.delete()

async def process_report_request(event, start_date, end_date, period_name_template):
    """Helper to reduce boilerplate in report commands."""
    period_str = f"{start_date.strftime('%d-%b-%Y')}"
    if end_date and start_date != end_date:
        period_str += f" to {end_date.strftime('%d-%b-%Y')}"
    
    proc_msg = await event.reply(f"⏳ Getting invoices for {period_str}...")
    invoices = await get_invoices(await event.get_input_chat(), start_date, end_date)
    await proc_msg.delete()
    await send_report(event, invoices, period_name_template.format(start=start_date, end=end_date))

# ====== AUTO-REPLY HANDLERS ======
@client.on(events.NewMessage(pattern=r'(?i)^(/hi bot|/ជំរាបសួរបង|/សួស្តីបង|/hi|/hello|0)$'))
async def send_auto_menu(event):
    """Handles various greetings to trigger the auto-reply menu."""
    # This also handles the case where a user might just send a sticker
    # by being registered with a broader pattern and checking inside.
    try:
        # Send sticker
        sticker_msg = await client.get_messages(SOURCE_CHANNEL, ids=59)
        if sticker_msg and sticker_msg.sticker:
            await client.send_file(event.chat_id, sticker_msg.sticker)

        # Send menu message
        msg = await client.get_messages(SOURCE_CHANNEL, ids=MENU_MSG_ID)
        await client.send_message(event.chat_id, msg)
    except Exception as e:
        await event.reply("⚠️ មិនអាចទាញយកម៉ឺនុយបានទេ។")
        logging.error(f"Menu Error: {e}")

@client.on(events.NewMessage(func=lambda e: e.sticker and not e.out and not e.text))
async def handle_sticker_greeting(event):
    """Specifically handles incoming stickers to trigger the menu."""
    await send_auto_menu(event)

@client.on(events.NewMessage(pattern=r'^[1-7]$'))
async def handle_text_selection(event):
    """Handles numeric selection for the auto-reply menu."""
    selection = event.raw_text.strip()
    await send_content_by_selection(event, selection)

# ====== INVOICE BOT HANDLERS ======
@client.on(events.NewMessage(pattern=r"^/menu$"))
async def invoice_menu(event):
    menu_text = (
        "🛠 Invoice Bot Menu\n\n"
        "Basic Commands:\n"
        " /today - Today's invoices\n"
        " /yesterday - Yesterday's invoices\n"
        " /this_month - This month invoices\n"
        " /last_month - Last month invoices\n"
        " /week - Last 7 days invoices\n\n"
        "Advanced Commands:\n"
        " /range DD-MM-YYYY DD-MM-YYYY - Custom date range\n"
        " /find <invoice_no> - Find invoice by number\n"
        " /stats - Monthly statistics"
    )
    await event.reply(menu_text)

@client.on(events.NewMessage(pattern=r"^/yesterday$"))
async def yesterday_report(event):
    target_date = datetime.now().date() - timedelta(days=1)
    await process_report_request(event, target_date, None, "Yesterday ({start:%d-%b-%Y})")

@client.on(events.NewMessage(pattern=r"^/today$"))
async def today_report(event):
    target_date = datetime.now().date()
    await process_report_request(event, target_date, None, "Today ({start:%d-%b-%Y})")

@client.on(events.NewMessage(pattern=r"^/this_month$"))
async def this_month_report(event):
    today = datetime.now().date()
    first_day = today.replace(day=1)
    await process_report_request(event, first_day, today, "This Month ({end:%b-%Y})")

@client.on(events.NewMessage(pattern=r"^/last_month$"))
async def last_month_report(event):
    today = datetime.now().date()
    first_day = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    last_day = today.replace(day=1) - timedelta(days=1)
    await process_report_request(event, first_day, last_day, "Last Month ({start:%b-%Y})")

@client.on(events.NewMessage(pattern=r"^/week$"))
async def week_report(event):
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=7)
    await process_report_request(event, start_date, end_date, "Weekly Report ({start:%d-%b-%Y} to {end:%d-%b-%Y})")

@client.on(events.NewMessage(pattern=r'^/range (\d{2}-\d{2}-\d{4}) (\d{2}-\d{2}-\d{4})$'))
async def range_report(event):
    try:
        start = datetime.strptime(event.pattern_match.group(1), '%d-%m-%Y').date()
        end = datetime.strptime(event.pattern_match.group(2), '%d-%m-%Y').date()
        if start > end:
            await event.reply("⚠️ Error: Start date must be before end date")
            return
        await process_report_request(event, start, end, "Date Range: {start:%d-%b-%Y} to {end:%d-%b-%Y}")
    except ValueError:
        await event.reply("⚠️ Invalid date format. Use: /range DD-MM-YYYY DD-MM-YYYY")

@client.on(events.NewMessage(pattern=r'^/find (\d+)$'))
async def find_report(event):
    target_no = event.pattern_match.group(1).zfill(6)
    proc_msg = await event.reply(f"🔍 Searching for invoice #{target_no}...")
    invoices = await get_invoices(await event.get_input_chat(), datetime(2000,1,1).date(), datetime.now().date())
    
    found = [i for i in invoices if i['no'].zfill(6) == target_no]
    await proc_msg.delete()
    if found:
        await send_report(event, found, f"Invoice #{target_no} Details")
    else:
        await event.reply(f"❌ Invoice #{target_no} not found in records")

# ====== MAIN EXECUTION ======
async def main():
    print("🔄 Connecting to Telegram...")
    try:
        # Start a dummy web server if PORT is defined (Required for Render Web Services)
        if "PORT" in os.environ:
            from aiohttp import web
            async def health_check(request):
                return web.Response(text="Bot is running!")
            
            app = web.Application()
            app.router.add_get('/', health_check)
            runner = web.AppRunner(app)
            await runner.setup()
            site = web.TCPSite(runner, '0.0.0.0', int(os.environ["PORT"]))
            await site.start()
            print(f"🌍 Web server started on port {os.environ['PORT']}")

        # Smart Login: Prevents interactive prompts on Render
        if BOT_TOKEN:
            await client.start(bot_token=BOT_TOKEN)
        else:
            if "PORT" in os.environ and not SESSION_STRING:
                raise Exception("❌ CRITICAL: SESSION_STRING is missing on Render! Run gen_session.py locally first.")
            await client.start()
            
        async with client:
            me = await client.get_me()
            print(f"✅ Combined Bot is running as Userbot 👤: {me.first_name} (ID: {me.id})")
            print("Press Ctrl+C to stop.")
            await client.run_until_disconnected()
    except Exception as e:
        error_msg = f"🚨 **BOT CRASHED**\n\nReason: `{str(e)}`"
        print(f"\n{error_msg}")
        try:
            if not client.is_connected(): await client.connect()
            await client.send_message('me', error_msg)
        except:
            print("❌ Failed to send alert to Telegram.")
        raise e

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n🛑 Bot stopped by user.")
    except Exception as e:
        print(f"\n❌ Error: {e}")
